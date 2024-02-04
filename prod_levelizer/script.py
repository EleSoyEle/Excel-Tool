import  openpyxl
import numpy as np
import matplotlib.pyplot as plt

sheet = openpyxl.load_workbook("prodlevelizer.xlsx")
ws = sheet.active

hoy_row = 4 #Fila hoy

ajuste1 = []
ajuste2 = []
entradas_ajustadas = []
inventario_forecast = []
inventario = ws.cell(row=hoy_row-1,column=5).value

def calcular_dia(i,columna_hoy):
    min_inv = ws.cell(row=columna_hoy+i-1,column=6).value
    max_inv = ws.cell(row=columna_hoy+i-1,column=7).value
    entradas_produccion = ws.cell(row=columna_hoy+i-1,column=4).value
    salidas_produccion = ws.cell(row=columna_hoy+i-1,column=3).value
    
    #Calculamos el ajuste1
    if i==0:
        if inventario>min_inv and inventario<max_inv:
            ajuste1.append(1)
        elif inventario>max_inv:
            ajuste1.append(0.9)
        else:
            ajuste1.append(1.1)
    #Calculamos el ajuste2,la primer entrada ajustada y el primer forecast
    elif i==1:
        entradas_ajustadas.append(ajuste1[-1]*entradas_produccion)
        finventario = inventario+entradas_ajustadas[-1]-salidas_produccion
        inventario_forecast.append(finventario)
        
        if finventario>min_inv and finventario<max_inv:
            ajuste2.append(1)
        elif finventario>max_inv:
            ajuste2.append(0.9)
        else:
            ajuste2.append(1.1)
    #Continuamos con el punto anterior
    elif i>1:
        entradas_ajustadas.append(ajuste2[-1]*entradas_produccion)
        finventario_anterior = inventario_forecast[-1]
        finventario = finventario_anterior+entradas_ajustadas[-1]-salidas_produccion
        inventario_forecast.append(finventario)
        if finventario>min_inv and finventario<max_inv:
            ajuste2.append(1)
        elif finventario>max_inv:
            ajuste2.append(0.9)
        else:
            ajuste2.append(1.1)
def calcular_semana(columna_hoy):
    for i in range(7):
        calcular_dia(i,columna_hoy)