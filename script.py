import  openpyxl
import numpy as np
import matplotlib.pyplot as plt

sheet = openpyxl.load_workbook("prodlevelizer.xlsx")
ws = sheet.active

hoy_row = 4 #Fila hoy


def calcular_dia(i,fila_hoy,ajuste1,ajuste2,entradas_ajustadas,inventario_forecast,inventario):

    min_inv = ws.cell(row=fila_hoy+i-1,column=6).value
    max_inv = ws.cell(row=fila_hoy+i-1,column=7).value
    entradas_produccion = ws.cell(row=fila_hoy+i-1,column=4).value
    salidas_produccion = ws.cell(row=fila_hoy+i-1,column=3).value
    
    #Calculamos el ajuste1
    if i==0:
        if inventario>min_inv and inventario<max_inv:
            ajuste1.append(1)
        elif inventario>max_inv:
            ajuste1.append(0.9)
        else:
            ajuste1.append(1.1)
        ws.cell(row=hoy_row-1,column=8).value = ajuste1[-1]
    #Calculamos el ajuste2,la primer entrada ajustada y el primer forecast
    elif i==1:
        print(ajuste1)
        entrada_ajustada = ajuste1[-1]*entradas_produccion
        entradas_ajustadas.append(entrada_ajustada)
        ws.cell(row=hoy_row+i-1,column=9).value = entrada_ajustada
        finventario = inventario+entradas_ajustadas[-1]-salidas_produccion
        inventario_forecast.append(finventario)
        ws.cell(row=hoy_row+i-1,column=10).value = finventario
        if finventario>min_inv and finventario<max_inv:
            ajuste2.append(1)
        elif finventario>max_inv:
            ajuste2.append(0.9)
        else:
            ajuste2.append(1.1)
        ws.cell(row=hoy_row+i-1,column=11).value = ajuste2[-1]
    #Continuamos con el punto anterior
    elif i>1:
        entrada_ajustada = ajuste2[-1]*entradas_produccion
        ws.cell(row=hoy_row+i-1,column=9).value = entrada_ajustada
        entradas_ajustadas.append(entrada_ajustada)
        finventario_anterior = inventario_forecast[-1]
        finventario = finventario_anterior+entradas_ajustadas[-1]-salidas_produccion
        inventario_forecast.append(finventario)
        ws.cell(row=hoy_row+i-1,column=10).value = finventario

        if finventario>min_inv and finventario<max_inv:
            ajuste2.append(1)
        elif finventario>max_inv:
            ajuste2.append(0.9)
        else:
            ajuste2.append(1.1)
        ws.cell(row=hoy_row+i-1,column=11).value = ajuste2[-1]
def calcular_semana(fila_hoy):
    ajuste1 = []
    ajuste2 = []
    entradas_ajustadas = []
    inventario_forecast = []
    inventario = ws.cell(row=hoy_row-1,column=5).value
    classes = ["Ajuste 1","Entradas ajustadas","Inventario forecast","Ajuste2"]
    style = ws.cell(row=1,column=1)._style
    for i in range(len(classes)):
        ws.cell(row=1,column=i+8).value = classes[i]
        ws.cell(row=1,column=i+8)._style = style
        print(ws.cell(row=1,column=i+8).value)
        #ws.cell(row=1,column=i+1).font = openpyxl.styles.Font(b=True, color="00FFFF00")
    for i in range(7):
        calcular_dia(i,fila_hoy,ajuste1,ajuste2,entradas_ajustadas,inventario_forecast,inventario)


calcular_semana(4)
sheet.save("nuevo_levelizer.xlsx")